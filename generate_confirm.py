import re
import shutil
import time
import win32com.client
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Place "Confirmation Workbook.xlsx" in the same directory as this script
TEMPLATE_PATH = Path(__file__).parent / "Confirmation Workbook.xlsx"
OUTPUT_DIR = Path(__file__).parent / "confirmations"

# Maps Smartsheet "Index" column values to the contract price type shown on the confirmation
INDEX_MAP = {
    "tetco m2 if":           "Texas Eastern M2 Receipts IF",
    "tetco m2 gda":          "Texas Eastern M2 Receipts GDA",
    "tetco m2 gd":           "Texas Eastern M2 Receipts GDA",
    "eastern gas gd":        "Eastern Gas South GDA",
    "eastern gas gda":       "Eastern Gas South GDA",
    "eastern gas south gda": "Eastern Gas South GDA",
    "eastern gas if":        "Eastern Gas South IF",
    "eastern gas south if":  "Eastern Gas South IF",
}


def find_sheet_name(counterparty: str, direction: str) -> str | None:
    """Return the matching sheet name by checking against the template workbook."""
    wb = load_workbook(TEMPLATE_PATH, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    is_purchase = direction.lower() in ("buy", "purchase")
    cp_lower = counterparty.lower()
    best_match, best_score = None, 0
    for sheet_name in sheet_names:
        if sheet_name == "Sheet1":
            continue
        sheet_is_purchase = sheet_name.lower().endswith("(purchase)")
        if sheet_is_purchase != is_purchase:
            continue
        base = sheet_name.lower().replace(" (purchase)", "").strip()
        if base in cp_lower and len(base) > best_score:
            best_score = len(base)
            best_match = sheet_name
    return best_match


def parse_price(row: dict) -> tuple[str, float]:
    """Return (price_type_label, price_value) from a Smartsheet row."""
    fixed = row.get("Fixed Price")
    if fixed and str(fixed).strip() not in ("", "None", "-"):
        try:
            val = float(str(fixed).replace("$", "").replace(",", "").strip())
            return "Fixed Price", val
        except ValueError:
            pass
    index = str(row.get("Index", "")).strip()
    price_type = INDEX_MAP.get(index.lower(), index)
    diff_str = str(row.get("Differential", "0")).replace("$", "").replace(",", "").strip()
    try:
        price_val = float(diff_str)
    except (ValueError, TypeError):
        price_val = 0.0
    return price_type, price_val


def parse_volume(vol_str) -> int:
    if not vol_str:
        return 0
    try:
        return int(str(vol_str).replace(",", ""))
    except ValueError:
        return 0


def parse_date(date_str) -> datetime | None:
    if not date_str:
        return None
    s = str(date_str)[:10]
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def generate_confirm(row: dict) -> Path | None:
    """
    Open the template workbook in Excel, fill in trade data via COM (preserving
    all images/formatting), export the target sheet as PDF, then close without saving.

    Cell mapping (Excel row, col — 1-indexed):
        (4,  13) – Confirmation date
        (24,  5) – Delivery Period begin
        (24,  9) – Delivery Period end
        (26,  4) – Contract Price type label
        (26,  7) – Contract Price value
        (28,  4) – Firm Qty (MMBtu/d)
        (30,  4) – Delivery Point
        (48,  9) – Seller signature date  (sale sheets,     Vickery = Seller)
        (48,  3) – Buyer  signature date  (purchase sheets, Vickery = Buyer)
    """
    counterparty = row.get("Counterparty", "Unknown")
    direction    = row.get("Direction", "Sell")
    trade_id     = row.get("Trade ID", "0")

    if not TEMPLATE_PATH.exists():
        print(f"  [ERROR] Confirmation Workbook not found at: {TEMPLATE_PATH}")
        return None

    sheet_name = find_sheet_name(counterparty, direction)
    if not sheet_name:
        print(f"  [WARN ] No template sheet for '{counterparty}' ({direction}) – skipping")
        return None

    trade_date     = parse_date(row.get("Trade Date"))
    begin_date     = parse_date(row.get("Trade Begin Date"))
    end_date       = parse_date(row.get("Trade End Date"))
    volume         = parse_volume(row.get("Volume (MMBtu/d)"))
    delivery_point = row.get("Point", "")
    price_type, price_val = parse_price(row)
    is_purchase    = direction.lower() in ("buy", "purchase")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    safe_cp   = re.sub(r"[^\w]", "_", counterparty)[:25].strip("_")
    date_tag  = trade_date.strftime("%Y%m%d") if trade_date else "nodate"
    temp_path = OUTPUT_DIR / f"_tmp_{trade_id}_{safe_cp}_{date_tag}.xlsx"
    pdf_path  = OUTPUT_DIR / f"{trade_id}_{safe_cp}_{date_tag}.pdf"

    # Remove any leftover files from a previous failed run
    temp_path.unlink(missing_ok=True)
    pdf_path.unlink(missing_ok=True)

    # Copy the template so Excel has its own file to work with (preserves all images)
    shutil.copy2(TEMPLATE_PATH, temp_path)

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wbk = excel.Workbooks.Open(str(temp_path.resolve()))
        time.sleep(2)  # let Excel finish loading all 34 sheets before exporting
        ws  = wbk.Sheets(sheet_name)
        ws.Activate()

        # Fill in all variable fields directly in Excel (images stay intact)
        try:
            confirm_id = 10**9 + int(str(trade_id).replace(",", ""))
        except (ValueError, TypeError):
            confirm_id = 10**9
        ws.Cells(6, 13).Value = confirm_id

        if trade_date:
            date_str = trade_date.strftime("%m/%d/%Y")
            ws.Cells(4,  13).Value = date_str
            ws.Cells(48, 3 if is_purchase else 9).Value = date_str
        if begin_date:
            ws.Cells(24, 5).Value = begin_date.strftime("%m/%d/%Y")
        if end_date:
            ws.Cells(24, 9).Value = end_date.strftime("%m/%d/%Y")
        ws.Cells(26, 4).Value = price_type
        ws.Cells(26, 7).Value = price_val
        ws.Cells(28, 4).Value = volume
        ws.Cells(30, 4).Value = delivery_point

        ws.ExportAsFixedFormat(0, str(pdf_path.resolve()))
        wbk.Close(False)  # discard — original template is never touched
    finally:
        excel.Quit()

    temp_path.unlink(missing_ok=True)

    print(f"  [OK] PDF saved: {pdf_path.name}")
    return pdf_path
